"""Build the two-sheet relabelling workbook from the flat per-image spreadsheet.

The input is either the original relabelling .xlsx or the dataset builder's dataset.csv;
farms already covered by a workbook in --exclude-labelled are dropped.

Sheet 1 ("Farm labels"): one row per Farm PFI, with the original label columns and a
fixed-option confidence dropdown (High/Medium/Low) next to each label.

Sheet 2 ("Image labels"): one row per image, as the input sheet stands, with the
one-hot label columns replaced by a single dropdown holding the existing classes plus
Multiple Classes / Paddock / Other/Industrial / Ambiguous.

Both sheets reuse the styling of the input workbook: Arial 10, dark slate header band,
alternating shading per farm, thin gridline borders with a medium rule separating the
key columns from the label columns and closing off each farm.
"""

import argparse
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_INPUT = Path(
    "/home/mannixe/FLIP/flip-geoimage-dataset-builder/"
    "original_new_2026_07_24_generalisation/dataset.csv"
)
DEFAULT_OUTPUT = Path("output/2026_07_24_generalisation_relabel_farm_and_image_nsw.xlsx")
DEFAULT_LABELLED = Path("labelled_sheets")

# Columns in the input sheet that describe the farm/image rather than the label.
KEY_COLUMNS = [
    "source",
    "Farm PFI",
    "source_image_path",
    "Farm_type (previous)",
]
IMAGE_COLUMN = "image_path"
COMMENTS_COLUMN = "Comments"
COUNT_COLUMN = "n_images"
IMAGE_LABEL_COLUMN = "Label"
PFI_COLUMN = "Farm PFI"

# Only these reaches are kept; matched case-insensitively against the start of `source`.
# The NSW case studies, as opposed to the Victorian ones (bacchusmarsh/balliang/
# gisborne/wyuna).
DEFAULT_SOURCES = ["bega", "caniaba", "freemans", "mangrove", "nowra"]

# dataset.csv carries its labels as binary_* one-hots; the workbook instead uses the
# class list of the earlier relabelling sheets, so the two sets of labels concatenate.
CSV_LABELS = [
    "aqua",
    "commercialpig",
    "freerangepig",
    "backyardpig",
    "sheep",
    "beef",
    "dairy",
    "goat",
    "horse",
    "poultry",
    "residential",
]
CSV_COLUMNS = {
    "source": "source",
    PFI_COLUMN: "PFI",
    "source_image_path": "source_image_path",
    "Farm_type (previous)": "Farm_type",
    IMAGE_COLUMN: "image_path",
}
# Widths taken from the original relabelling workbook; the rest keep Excel's default.
CSV_WIDTHS = {
    "source": 24.0,
    PFI_COLUMN: 34.0,
    "source_image_path": 55.28,
    "Farm_type (previous)": 20.0,
    IMAGE_COLUMN: 122.29,
    "aqua": 11.0,
    "commercialpig": 19.57,
    "freerangepig": 16.86,
    "backyardpig": 17.71,
    "sheep": 11.0,
    "poultry": 12.57,
    "residential": 14.86,
    COMMENTS_COLUMN: 34.0,
}

PRESENT_OPTIONS = ["X"]
CONFIDENCE_OPTIONS = ["High", "Medium", "Low"]
EXTRA_IMAGE_OPTIONS = [
    "Multiple Classes",
    "Paddock",
    "Other/Industrial",
    "Ambiguous",
]

FARM_SHEET = "Farm labels"
IMAGE_SHEET = "Image labels"
LISTS_SHEET = "Lists"

# Styling copied from the input workbook.
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF44546A")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_HEIGHT = 30

BODY_FONT = Font(name="Arial", size=10, color="FF000000")
TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
LABEL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
# Farms alternate between these two fills, starting with the tinted one.
BAND_FILLS = [
    PatternFill("solid", fgColor="FFE8EAED"),
    PatternFill("solid", fgColor="FFFFFFFF"),
]

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
DEFAULT_WIDTH = 11
CONFIDENCE_WIDTH = 12
LABEL_WIDTH = 19.57


def border(left_medium: bool = False, right_medium: bool = False, bottom_medium: bool = False) -> Border:
    """Thin gridlines, with medium rules marking column groups and farm boundaries."""
    return Border(
        left=MEDIUM if left_medium else THIN,
        right=MEDIUM if right_medium else THIN,
        top=THIN,
        bottom=MEDIUM if bottom_medium else THIN,
    )


def read_source(path: Path) -> tuple[list[str], list[dict], dict[str, float]]:
    """Read the input into a header list, row dicts, and per-column widths."""
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_workbook(path)


def read_csv(path: Path) -> tuple[list[str], list[dict], dict[str, float]]:
    """Read dataset.csv, renaming its columns onto the relabelling sheet's schema.

    The label columns are left blank: this workbook is what produces those labels.
    """
    header = KEY_COLUMNS + [IMAGE_COLUMN] + CSV_LABELS + [COMMENTS_COLUMN]
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = [
            dict.fromkeys(header) | {name: source_row[column] for name, column in CSV_COLUMNS.items()}
            for source_row in csv.DictReader(handle)
        ]
    for row in rows:
        pfi = row[PFI_COLUMN]
        row[PFI_COLUMN] = int(pfi) if str(pfi).isdigit() else pfi
        # Farms the builder left unlabelled read as an empty cell rather than blank.
        row["Farm_type (previous)"] = row["Farm_type (previous)"] or "NA"
    return header, rows, CSV_WIDTHS


def read_workbook(path: Path) -> tuple[list[str], list[dict], dict[str, float]]:
    """Read the input sheet into a header list, row dicts, and per-column widths."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    row_iter = sheet.iter_rows(values_only=True)
    header = [cell for cell in next(row_iter) if cell is not None]
    rows = [
        dict(zip(header, values))
        for values in row_iter
        if any(value is not None for value in values)
    ]
    # Only the columns the input sets explicitly; the rest keep Excel's default width.
    dimensions = sheet.column_dimensions
    widths = {
        name: dimensions[letter].width
        for index, name in enumerate(header, start=1)
        if (letter := get_column_letter(index)) in dimensions
    }
    workbook.close()
    return header, rows, widths


def filter_sources(rows: list[dict], sources: list[str]) -> list[dict]:
    """Keep only rows whose `source` shapefile starts with one of the named reaches."""
    wanted = tuple(source.lower() for source in sources)
    return [row for row in rows if str(row["source"]).lower().startswith(wanted)]


def labelled_pfis(directory: Path) -> set[str]:
    """Every Farm PFI appearing in a workbook under `directory`, as strings."""
    pfis: set[str] = set()
    for path in sorted(directory.glob("*.xlsx")):
        if path.name.startswith("~$"):  # Excel's lock files
            continue
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None) or ()
            if PFI_COLUMN not in header:
                continue
            index = header.index(PFI_COLUMN)
            pfis.update(str(values[index]) for values in rows if values[index] is not None)
        workbook.close()
    return pfis


def drop_labelled(rows: list[dict], pfis: set[str]) -> list[dict]:
    """Drop every image of a farm that has already been through a labelling sheet."""
    return [row for row in rows if str(row[PFI_COLUMN]) not in pfis]


def label_columns(header: list[str]) -> list[str]:
    """The class columns: everything between image_path and Comments."""
    start = header.index(IMAGE_COLUMN) + 1
    end = header.index(COMMENTS_COLUMN)
    return header[start:end]


def group_by_farm(rows: list[dict]) -> list[dict]:
    """One record per Farm PFI, keeping first-seen order and image counts."""
    farms: dict[object, dict] = {}
    for row in rows:
        pfi = row["Farm PFI"]
        farm = farms.get(pfi)
        if farm is None:
            farm = {key: row[key] for key in KEY_COLUMNS}
            farm[COUNT_COLUMN] = 0
            farms[pfi] = farm
        farm[COUNT_COLUMN] += 1
    return list(farms.values())


def write_header(sheet, header: list[str], medium_left: set[int]) -> None:
    sheet.append(header)
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = border(
            left_medium=index in medium_left,
            right_medium=index == len(header),
        )
    sheet.row_dimensions[1].height = HEADER_HEIGHT


def style_row(
    sheet,
    row: int,
    n_columns: int,
    centred: set[int],
    medium_left: set[int],
    fill: PatternFill,
    bottom_medium: bool,
) -> None:
    for index in range(1, n_columns + 1):
        cell = sheet.cell(row=row, column=index)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = LABEL_ALIGNMENT if index in centred else TEXT_ALIGNMENT
        cell.border = border(
            left_medium=index in medium_left,
            right_medium=index == n_columns,
            bottom_medium=bottom_medium,
        )


def set_widths(sheet, widths: list[float | None]) -> None:
    for index, width in enumerate(widths, start=1):
        if width is not None:
            sheet.column_dimensions[get_column_letter(index)].width = width


def finish_sheet(sheet, n_columns: int, last_row: int) -> None:
    sheet.freeze_panes = f"{get_column_letter(len(KEY_COLUMNS) + 2)}2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(n_columns)}{last_row}"


def add_lists_sheet(workbook, image_options: list[str]):
    """Hidden sheet holding the dropdown options, referenced by the validations."""
    sheet = workbook.create_sheet(LISTS_SHEET)
    columns = {
        "Confidence": CONFIDENCE_OPTIONS,
        "Image label": image_options,
        "Class present": PRESENT_OPTIONS,
    }
    sheet.append(list(columns))
    for index, options in enumerate(columns.values(), start=1):
        for offset, value in enumerate(options, start=2):
            sheet.cell(row=offset, column=index, value=value)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = TEXT_ALIGNMENT
    set_widths(sheet, [16, 22, 14])
    sheet.sheet_state = "hidden"
    return sheet


def list_validation(formula_range: str, prompt: str) -> DataValidation:
    validation = DataValidation(
        type="list",
        formula1=f"={LISTS_SHEET}!{formula_range}",
        allow_blank=True,
        showDropDown=False,  # False means "show the in-cell dropdown arrow"
        showInputMessage=True,
        showErrorMessage=True,
    )
    validation.error = "Pick one of the listed options."
    validation.errorTitle = "Invalid entry"
    validation.prompt = prompt
    validation.promptTitle = "Select a value"
    return validation


def build_farm_sheet(workbook, farms: list[dict], labels: list[str], widths: dict[str, float]) -> None:
    """One row per farm; each label column is followed by its confidence dropdown."""
    sheet = workbook.create_sheet(FARM_SHEET)
    header = KEY_COLUMNS + [COUNT_COLUMN]
    for label in labels:
        header += [label, f"{label} confidence"]
    header.append(COMMENTS_COLUMN)

    first_label = len(KEY_COLUMNS) + 2
    label_columns_ = {first_label + 2 * index: label for index, label in enumerate(labels)}
    confidence_columns = [column + 1 for column in label_columns_]
    # A medium rule opens the label block and separates each label/confidence pair.
    medium_left = set(label_columns_)
    centred = {len(KEY_COLUMNS) + 1, *label_columns_, *confidence_columns}

    write_header(sheet, header, medium_left)
    for index, farm in enumerate(farms):
        sheet.append([farm[key] for key in KEY_COLUMNS] + [farm[COUNT_COLUMN]])
        style_row(
            sheet,
            sheet.max_row,
            len(header),
            centred,
            medium_left,
            BAND_FILLS[index % 2],
            bottom_medium=index == len(farms) - 1,
        )

    last_row = sheet.max_row
    present = list_validation(
        f"$C$2:$C${1 + len(PRESENT_OPTIONS)}",
        "X if this class is present on the farm, otherwise leave blank.",
    )
    confidence = list_validation(
        f"$A$2:$A${1 + len(CONFIDENCE_OPTIONS)}",
        "How confident are you in this farm-level label?",
    )
    for validation, columns in ((present, label_columns_), (confidence, confidence_columns)):
        sheet.add_data_validation(validation)
        for column in columns:
            letter = get_column_letter(column)
            validation.add(f"{letter}2:{letter}{last_row}")

    column_widths = [widths.get(key) for key in KEY_COLUMNS] + [DEFAULT_WIDTH]
    for label in labels:
        column_widths += [widths.get(label), CONFIDENCE_WIDTH]
    column_widths.append(widths.get(COMMENTS_COLUMN))
    set_widths(sheet, column_widths)
    finish_sheet(sheet, len(header), last_row)


def build_image_sheet(workbook, rows: list[dict], image_options: list[str], widths: dict[str, float]) -> None:
    """One row per image, with a single class dropdown replacing the one-hot columns."""
    sheet = workbook.create_sheet(IMAGE_SHEET)
    header = KEY_COLUMNS + [IMAGE_COLUMN, IMAGE_LABEL_COLUMN, COMMENTS_COLUMN]
    label_column = len(KEY_COLUMNS) + 2
    medium_left = {label_column}
    centred = {label_column}

    write_header(sheet, header, medium_left)
    band = 0
    for index, row in enumerate(rows):
        sheet.append(
            [row[key] for key in KEY_COLUMNS]
            + [row[IMAGE_COLUMN], None, row.get(COMMENTS_COLUMN)]
        )
        # Shading alternates per farm, and each farm is closed off with a medium rule.
        last_of_farm = (
            index == len(rows) - 1 or rows[index + 1]["Farm PFI"] != row["Farm PFI"]
        )
        style_row(
            sheet,
            sheet.max_row,
            len(header),
            centred,
            medium_left,
            BAND_FILLS[band % 2],
            bottom_medium=last_of_farm,
        )
        band += last_of_farm

    last_row = sheet.max_row
    validation = list_validation(
        f"$B$2:$B${1 + len(image_options)}",
        "Pick the class shown in this image.",
    )
    sheet.add_data_validation(validation)
    letter = get_column_letter(label_column)
    validation.add(f"{letter}2:{letter}{last_row}")

    set_widths(
        sheet,
        [widths.get(key) for key in KEY_COLUMNS]
        + [widths.get(IMAGE_COLUMN), LABEL_WIDTH, widths.get(COMMENTS_COLUMN)],
    )
    finish_sheet(sheet, len(header), last_row)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--sources",
        nargs="*",
        default=DEFAULT_SOURCES,
        help="reaches to keep, matched against the start of the source shapefile name "
        "(pass with no values to keep everything)",
    )
    parser.add_argument(
        "--exclude-labelled",
        type=Path,
        default=DEFAULT_LABELLED,
        help="directory of completed labelling workbooks whose farms are already done "
        "(pass an empty string to keep everything)",
    )
    args = parser.parse_args()

    header, rows, widths = read_source(args.input)
    if args.sources:
        rows = filter_sources(rows, args.sources)
        if not rows:
            parser.error(f"no rows match sources {args.sources}")
    done = labelled_pfis(args.exclude_labelled) if args.exclude_labelled else set()
    if done:
        rows = drop_labelled(rows, done)
        if not rows:
            parser.error(f"every matching farm is already labelled in {args.exclude_labelled}")
    labels = label_columns(header)
    farms = group_by_farm(rows)
    image_options = labels + EXTRA_IMAGE_OPTIONS

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    add_lists_sheet(workbook, image_options)
    build_farm_sheet(workbook, farms, labels, widths)
    build_image_sheet(workbook, rows, image_options, widths)
    workbook.move_sheet(LISTS_SHEET, offset=2)
    workbook.active = 0
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)

    print(f"{args.input} -> {args.output}")
    print(f"  sources: {', '.join(sorted({row['source'] for row in rows}))}")
    print(f"  excluded: {len(done)} farms already labelled in {args.exclude_labelled}")
    print(f"  {FARM_SHEET}: {len(farms)} farms x {len(labels)} labels (+ confidence)")
    print(f"  {IMAGE_SHEET}: {len(rows)} images, {len(image_options)} dropdown options")


if __name__ == "__main__":
    main()
