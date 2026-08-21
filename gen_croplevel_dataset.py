"""Build a crop-level dataset from the labelling workbooks in labelled_sheets/.

The dataset builder's dataset.csv carries one row per crop but labels each row with
its *farm's* Farm_type, which is wrong for most crops: a farm's ten crops are usually
one shed and nine paddocks. The relabelling workbooks fix that — the "Image labels"
sheet holds one label per crop — so this script joins those labels back onto the
builder's rows and writes a dataset_relabelled.csv keyed on the crop label instead.
It is written alongside the build's own dataset.csv, never over it.

Joining is on (farm_uid, ecw_stem, building_cluster) rather than image_path: the
workbooks were written against the older PFI-keyed collection, so their image_path
still reads "<PFI>/buildings/<stem>_farm_<PFI>_building_<n>.tif" while the current
build is keyed on the farm UID. The triple is stable across both.

A "Multiple Classes" crop is genuinely multi-label — dairy and horse, poultry and
freerange pig, two real livestock classes in the one crop — and is resolved from the
classes the labeller named in its comment, or from its farm's X marks where the comment
was left blank. Such a crop carries every class it shows in binary_*, as the builder
already does for a farm labelled "dairy,beef". Only "Ambiguous" crops are dropped, and
outright: this rebuilds the dataset from scratch, so there is no historical split to
stay compatible with and no reason to write rows the loaders must know to skip.

Splits are geographic, so the held-out set is a different landscape rather than a
different farm in the same one:

    train/val  NSW  - bega, caniaba, freemans, mangrove, nowra
    test       VIC  - bacchusmarsh, balliang, gisborne, wyuna

Within NSW the train/val cut follows extract_imagery_aerial_csv.py: farm-grouped so a
farm's crops never straddle the two, stratified on the farm's class, VAL_FRACTION of
the farms, RANDOM_STATE for reproducibility.

Output is dataset_relabelled.csv plus relabelled_{train,val,test}_df.csv. They land in
the build's own directory by default, so the relative image_path values keep resolving
with no imagery moved; --output-dir elsewhere plus --imagery makes a standalone copy.
"""

import argparse
import collections
import os
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from sklearn.model_selection import train_test_split

DEFAULT_DATASET = Path(
    "/home/mannixe/FLIP/flip-geoimage-dataset-builder/"
    "original_new_2026_08_21_generalisation/dataset.csv"
)
DEFAULT_LABELLED = Path("labelled_sheets")

# Written alongside the builder's own dataset.csv / train_df.csv rather than over them:
# same directory, so the relative image_path values keep resolving without moving or
# copying any imagery, but under names that cannot collide with the build they came from.
DATASET_NAME = "dataset_relabelled.csv"
SPLIT_NAME = "relabelled_{split}_df.csv"

IMAGE_SHEET = "Image labels"
FARM_SHEET = "Farm labels"
UID_COLUMN = "Farm UID"
IMAGE_COLUMN = "image_path"
LABEL_COLUMN = "Label"
COMMENTS_COLUMN = "Comments"

# The reaches making up each region. Matched case-insensitively against the start of
# the `source` shapefile name, as in make_spreadsheet.py.
NSW_SOURCES = ["bega", "caniaba", "freemans", "mangrove", "nowra"]
VIC_SOURCES = ["bacchusmarsh", "balliang", "gisborne", "wyuna"]

# The workbook dropdowns offer the class list plus four catch-alls. Two of those are
# real crop classes — most crops in this collection are paddock or a non-farm building
# and a crop-level classifier has to be able to say so — and two say the labeller could
# not call it, which is not something to train on.
LABEL_RENAMES = {
    "goat": "goats",  # the workbooks say goat, the builder's binary_ column says goats
    "paddock": "paddock",
    "other/industrial": "other_industrial",
    # the labeller's free-text spellings, from the comments on "Multiple Classes" crops
    "freerange pigs": "freerangepig",
    "freerange pig": "freerangepig",
    "backyard pig": "backyardpig",
    "backyard pigs": "backyardpig",
    "commercial pig": "commercialpig",
    "commercial pigs": "commercialpig",
}
# "Multiple Classes" is not a class: the labeller names the classes they saw in the
# crop's Comments instead ("Poultry/Freerange pigs"), so the label is read from there
# and the crop comes through genuinely multi-label. These are not cases a most-specific
# rule could collapse — the pairs are dairy/horse, poultry/freerangepig,
# commercialpig/backyardpig, two real livestock classes in the one crop — so the crop
# carries both and it is binary_* that says so, exactly as the builder handles a farm
# labelled "dairy,beef". Only "Ambiguous" — the labeller could not call it at all — is
# left with nothing to train on.
MULTI_LABEL = "multiple classes"
COMMENT_SEPARATORS = "/,+&"
DEFAULT_DROP_LABELS = ["Ambiguous"]

# Every class a crop row may end up in; the binary_ one-hots are built from this list
# so a split's columns do not depend on which classes happen to appear in it.
CROP_CLASSES = [
    "aqua", "backyardpig", "beef", "commercialpig", "dairy", "freerangepig",
    "goats", "horse", "poultry", "residential", "sheep",
    "paddock", "other_industrial",
]
# The class columns of the "Farm labels" sheet, in the workbook's own spelling; each
# is followed by a "<class> confidence" column, which is not a label.
FARM_LABEL_COLUMNS = [
    "aqua", "commercialpig", "freerangepig", "backyardpig", "sheep", "beef",
    "dairy", "goat", "horse", "poultry", "residential",
]
# Not a farm class: excluded when picking the farm's stratum for the val split, and
# reported separately.
BACKGROUND_CLASSES = {"paddock", "other_industrial"}

VAL_FRACTION = 0.20
RANDOM_STATE = 42

# Carried through from the builder's dataset.csv. Its own processed_class/binary_*
# columns are farm-level and are replaced here, but the raw Farm_type is kept so the
# old and new labels can be compared.
CARRY_COLUMNS = [
    "image_path", "farm_uid", "ecw_stem", "building_cluster", "Farm_type", "PFI",
    "source", "filename", "Lat", "Long", "source_filenames", "source_image_path",
    "crop_width", "crop_height", "crop_res", "crop_std", "crop_fill_frac",
    "blank_reason",
]
JOIN_KEYS = ["farm_uid", "ecw_stem", "building_cluster"]


def crop_key(farm_uid, image_path):
    """The (farm_uid, ecw_stem, building_cluster) triple identifying one crop.

    Crop filenames read "{ecw_stem}_farm_{farm_key}_building_{cluster}.tif", where
    farm_key is whatever named the farm directory when the crop was written — a PFI
    on the collection the workbooks were built from, a farm UID now. The farm_key in
    the filename is therefore ignored; the UID column carries the current identity.
    """
    stem = Path(str(image_path)).stem
    ecw_stem, _, rest = stem.partition("_farm_")
    _, _, building = rest.partition("_building_")
    return str(farm_uid), ecw_stem, building


def read_sheet(path, name):
    """Rows of one sheet as dicts, or None if the workbook has no such sheet."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if name not in workbook.sheetnames:
        workbook.close()
        return None
    values = workbook[name].iter_rows(values_only=True)
    header = list(next(values, ()) or ())
    rows = [
        dict(zip(header, row))
        for row in values
        if any(value is not None for value in row)
    ]
    workbook.close()
    return rows


def workbooks(directory):
    return [p for p in sorted(directory.glob("*.xlsx")) if not p.name.startswith("~$")]


def read_crop_labels(directory):
    """Every labelled crop across the workbooks, keyed on its crop_key triple.

    A crop labelled in two workbooks (a reach relabelled after a first pass) keeps the
    label from the workbook read last, and is counted as a duplicate for the report.
    """
    labels = {}
    duplicates = []
    for path in workbooks(directory):
        rows = read_sheet(path, IMAGE_SHEET)
        if rows is None:
            continue
        for row in rows:
            label = row.get(LABEL_COLUMN)
            if label is None or not str(label).strip():
                continue
            key = crop_key(row[UID_COLUMN], row[IMAGE_COLUMN])
            if key in labels:
                duplicates.append(key)
            labels[key] = {
                "crop_label": str(label).strip(),
                "crop_comments": row.get(COMMENTS_COLUMN) or "",
                "label_workbook": path.name,
            }
    return labels, duplicates


def read_farm_labels(directory):
    """Per-farm class set from the "Farm labels" sheets, as a comma-separated string.

    The farm sheet marks each class the labeller saw anywhere on the farm with an X,
    so a farm can carry several. Kept alongside the crop label so a crop-level row
    still says what the whole farm is.
    """
    farms = collections.defaultdict(list)
    for path in workbooks(directory):
        rows = read_sheet(path, FARM_SHEET)
        if rows is None:
            continue
        for row in rows:
            uid = str(row.get(UID_COLUMN) or "")
            if not uid:
                continue
            marked = [
                normalise_label(name)
                for name in row
                if str(name) in FARM_LABEL_COLUMNS and row[name]
            ]
            for label in marked:
                if label not in farms[uid]:
                    farms[uid].append(label)
    return {uid: ",".join(labels) for uid, labels in farms.items()}


def normalise_label(label):
    """Workbook label -> dataset class name (lower-cased, catch-alls renamed)."""
    label = " ".join(str(label or "").split()).lower()
    return LABEL_RENAMES.get(label, label)


def comment_classes(comment):
    """The classes named in a "Multiple Classes" crop's comment.

    The labeller writes them free-text and separator-separated — "Poultry/Freerange
    pigs", "Dairy/Horse", "dairy/ backyard pig" — so each part is normalised through
    LABEL_RENAMES and anything that is not a known class is discarded. Returns [] for a
    comment naming none, which is how an unresolvable crop is detected.

    The result is sorted into CROP_CLASSES order, not the order it was written in: the
    same labeller wrote both "Dairy/Horse" and "Horse/Dairy" for the same pair, so the
    comment's own order carries no information and keeping it would only split one class
    combination across two spellings. The raw comment stays in crop_comments.
    """
    text = str(comment or "")
    for separator in COMMENT_SEPARATORS:
        text = text.replace(separator, "/")
    classes = {
        label for part in text.split("/")
        if (label := normalise_label(part)) in CROP_CLASSES
    }
    return sorted(classes, key=CROP_CLASSES.index)


def crop_classes(label, comment, farm_classes=()):
    """The class list for one crop, which for a multi-class crop has more than one entry.

    A "Multiple Classes" crop is resolved from its comment, falling back to the classes
    X-marked for its farm on the "Farm labels" sheet when the labeller left the comment
    blank. The fallback is only sound because the farm sheet is the same labeller
    recording what they saw on the same farm: the two balliang crops it resolves sit on
    farms marked commercialpig + freerangepig, which is the pair the crop shows.
    """
    label = normalise_label(label)
    if label == MULTI_LABEL:
        return comment_classes(comment) or sorted(
            {cls for cls in farm_classes if cls in CROP_CLASSES},
            key=CROP_CLASSES.index,
        )
    return [label] if label else []


def region_of(source, nsw_sources, vic_sources):
    """"nsw" / "vic" / "" for a source shapefile name."""
    name = str(source).lower()
    if name.startswith(tuple(s.lower() for s in nsw_sources)):
        return "nsw"
    if name.startswith(tuple(s.lower() for s in vic_sources)):
        return "vic"
    return ""


def farm_strata(df):
    """One class per farm, for stratifying the val split.

    A farm has many crop labels, so the farm's stratum is its most common non-background
    crop class — the thing the farm actually is — falling back to its most common label
    for a farm that is all paddock.
    """
    strata = {}
    for farm_uid, group in df.groupby("farm_uid")["processed_class"]:
        counts = group.value_counts()
        farm_classes = counts[~counts.index.isin(BACKGROUND_CLASSES)]
        strata[farm_uid] = (farm_classes if not farm_classes.empty else counts).index[0]
    return strata


def assign_val_split(df, val_fraction=VAL_FRACTION, random_state=RANDOM_STATE):
    """Promote ~val_fraction of the train farms to 'val'.

    Grouped on farm_uid so a farm's crops stay together, and stratified on the farm's
    class so the two sides stay comparable. Classes with fewer than two farms in the
    pool cannot be stratified and stay in train. Test rows are left untouched.
    """
    train = df[df["split"] == "train"]
    if train.empty:
        return df
    strata = farm_strata(train)
    farms = pd.DataFrame(
        {"farm_uid": list(strata), "stratum": list(strata.values())}
    )
    counts = farms["stratum"].value_counts()
    stratifiable = farms[farms["stratum"].isin(counts[counts >= 2].index)]
    if len(stratifiable) < 2:
        return df
    _, val_farms = train_test_split(
        stratifiable,
        test_size=val_fraction,
        stratify=stratifiable["stratum"],
        random_state=random_state,
    )
    df.loc[df["farm_uid"].isin(set(val_farms["farm_uid"])), "split"] = "val"
    return df


def build_dataset(dataset_csv, labelled_dir, nsw_sources, vic_sources,
                  drop_labels, val_fraction=VAL_FRACTION):
    """Join the crop labels onto the builder's rows and assign the splits."""
    source = pd.read_csv(dataset_csv, dtype={"building_cluster": str, "PFI": str})
    labels, duplicates = read_crop_labels(labelled_dir)
    farm_labels = read_farm_labels(labelled_dir)

    keys = [crop_key(row.farm_uid, row.image_path) for row in source.itertuples()]
    matched = [key in labels for key in keys]
    unmatched_sheet = set(labels) - set(keys)

    df = source.loc[matched, [c for c in CARRY_COLUMNS if c in source.columns]].copy()
    found = [labels[key] for key, keep in zip(keys, matched) if keep]
    df["crop_label"] = [f["crop_label"] for f in found]
    df["crop_comments"] = [f["crop_comments"] for f in found]
    df["label_workbook"] = [f["label_workbook"] for f in found]
    # the builder's farm-level label, kept for comparison against the crop label
    df["farm_processed_class"] = source.loc[matched, "processed_class"].values
    df["farm_labels"] = df["farm_uid"].map(farm_labels).fillna("")

    # a crop can carry more than one class ("Multiple Classes", resolved from its
    # comment), so the class list is the primary column and processed_class is its
    # first entry — the same relationship the builder's Farm_type/processed_class have
    classes = [
        crop_classes(label, comment, farm.split(",") if farm else ())
        for label, comment, farm in zip(
            df["crop_label"], df["crop_comments"], df["farm_labels"]
        )
    ]
    df["crop_classes"] = [",".join(c) for c in classes]
    df["processed_class"] = [c[0] if c else "" for c in classes]
    df["n_classes"] = [len(c) for c in classes]
    for cls in CROP_CLASSES:
        df["binary_" + cls] = [cls in c for c in classes]

    # "Ambiguous" says the labeller could not call the crop, and a multi-class crop
    # whose comment names no class cannot be resolved either; neither is something to
    # train or test on. This is a build from scratch, not an increment on a historical
    # split, so they are dropped outright rather than kept in dataset.csv under a split
    # the loaders would then have to know to skip.
    dropping = (
        df["processed_class"].isin({normalise_label(l) for l in drop_labels})
        | (df["n_classes"] == 0)
    )
    dropped = collections.Counter(
        f"{label} (neither the comment nor the farm sheet names a class)"
        if normalise_label(label) == MULTI_LABEL else label
        for label in df.loc[dropping, "crop_label"]
    )
    df = df[~dropping].copy()

    # geography decides the split
    df["region"] = df["source"].map(lambda s: region_of(s, nsw_sources, vic_sources))
    df["split"] = ""
    df.loc[df["region"] == "nsw", "split"] = "train"
    df.loc[df["region"] == "vic", "split"] = "test"
    df = assign_val_split(df, val_fraction)

    unknown = sorted(
        {cls for row in df["crop_classes"] for cls in row.split(",")} - set(CROP_CLASSES)
    )
    return df, unmatched_sheet, duplicates, unknown, dropped


def link_imagery(df, dataset_csv, output_dir, mode):
    """Put every crop (and its source image) under output_dir at its dataset path.

    The csv's image_path is relative to the dataset root, so linking the files in under
    the same relative path makes the output directory a dataset in its own right rather
    than a set of csvs pointing back at the build it came from.
    """
    root = Path(dataset_csv).parent
    if mode == "none" or root.resolve() == Path(output_dir).resolve():
        return 0, []
    paths = set(df["image_path"]) | {
        p for p in df["source_image_path"].fillna("") if p
    }
    linked, missing = 0, []
    for relative in sorted(paths):
        source = root / relative
        if not source.exists():
            missing.append(relative)
            continue
        target = output_dir / relative
        if target.exists() or target.is_symlink():
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        if mode == "copy":
            shutil.copy2(source, target)
        else:
            target.symlink_to(os.path.relpath(source.resolve(), target.parent))
        linked += 1
    return linked, missing


def report(df, unmatched_sheet, duplicates, unknown, dropped, args):
    print(f"{args.dataset} + {args.labelled}/ -> {args.output_dir}/{DATASET_NAME}")
    print(f"  {len(df)} labelled crops over {df['farm_uid'].nunique()} farms")
    for label, count in sorted(dropped.items()):
        print(f"  dropped {count} crops labelled {label!r}")
    if duplicates:
        print(f"  warning: {len(duplicates)} crops labelled in more than one workbook "
              "(the last workbook read wins)")
    if unmatched_sheet:
        print(f"  warning: {len(unmatched_sheet)} labelled crops are not in {args.dataset}")
        for key in sorted(unmatched_sheet)[:5]:
            print(f"    {key}")
    if unknown:
        print(f"  warning: labels not in CROP_CLASSES (add to LABEL_RENAMES or "
              f"--drop-labels): {unknown}")
    no_region = df[df["region"] == ""]
    if not no_region.empty:
        print(f"  warning: {len(no_region)} crops from sources in neither region: "
              f"{sorted(set(no_region['source']))}")

    multi = df[df["n_classes"] > 1]
    if not multi.empty:
        print(f"\n{len(multi)} multi-label crops over {multi['farm_uid'].nunique()} farms:")
        for combination, count in sorted(collections.Counter(multi["crop_classes"]).items()):
            print(f"  {count:>3}  {combination}")

    # a multi-label crop counts under each of its classes, so these are built from the
    # exploded class list rather than from processed_class alone
    exploded = (
        df.assign(cls=df["crop_classes"].str.split(","))
        .explode("cls")
        .reset_index(drop=True)
    )
    print("\ncrops per split x class:")
    print(pd.crosstab(exploded["cls"], exploded["split"]).to_string())
    print("\nunique farms per split x class:")
    print(
        exploded.pivot_table(
            index="cls", columns="split", values="farm_uid",
            aggfunc="nunique", fill_value=0,
        ).to_string()
    )
    print("\nunique farms per split:")
    print(df.groupby("split")["farm_uid"].nunique().to_string())
    print(f"total unique farms: {df['farm_uid'].nunique()}")
    # a farm's crops must stay on one side of the split
    straddling = df.groupby("farm_uid")["split"].nunique()
    straddling = straddling[straddling > 1]
    if not straddling.empty:
        print(f"  warning: {len(straddling)} farms straddle splits: "
              f"{sorted(straddling.index)[:5]}")

    print("\ncrop label vs the builder's farm-level label:")
    print(pd.crosstab(df["processed_class"], df["farm_processed_class"]).to_string())


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dataset", type=Path, default=DEFAULT_DATASET,
                        help="the dataset builder's dataset.csv to relabel")
    parser.add_argument("--labelled", type=Path, default=DEFAULT_LABELLED,
                        help="directory of completed labelling workbooks")
    parser.add_argument("--output-dir", type=Path, default=None,
                        help=f"where to write {DATASET_NAME} and the per-split csvs "
                             "(default: alongside --dataset, so image_path resolves as-is)")
    parser.add_argument("--nsw-sources", nargs="*", default=NSW_SOURCES,
                        help="reaches making up the train/val region")
    parser.add_argument("--vic-sources", nargs="*", default=VIC_SOURCES,
                        help="reaches making up the test region")
    parser.add_argument("--drop-labels", nargs="*", default=DEFAULT_DROP_LABELS,
                        help="workbook labels dropped from the build entirely "
                             "(pass with no values to keep everything)")
    parser.add_argument("--val-fraction", type=float, default=VAL_FRACTION,
                        help=f"share of the train farms promoted to val (default: {VAL_FRACTION})")
    parser.add_argument("--imagery", choices=["symlink", "copy", "none"], default="none",
                        help="link/copy the crops under --output-dir, making it a "
                             "standalone dataset. Only useful with --output-dir "
                             "elsewhere; writing beside --dataset needs nothing "
                             "(default: none)")
    args = parser.parse_args()
    if args.output_dir is None:
        args.output_dir = args.dataset.parent

    df, unmatched_sheet, duplicates, unknown, dropped = build_dataset(
        args.dataset, args.labelled, args.nsw_sources, args.vic_sources,
        args.drop_labels, args.val_fraction,
    )
    if df.empty:
        parser.error(f"no crop in {args.dataset} matched a label in {args.labelled}")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    columns = (
        [c for c in CARRY_COLUMNS if c in df.columns]
        + ["region", "crop_label", "crop_classes", "n_classes", "processed_class",
           "farm_processed_class", "farm_labels", "crop_comments", "label_workbook"]
        + ["binary_" + cls for cls in CROP_CLASSES]
        + ["split"]
    )
    df = df[columns]
    df.to_csv(args.output_dir / DATASET_NAME, index=False)

    report(df, unmatched_sheet, duplicates, unknown, dropped, args)
    print()
    for name in ("train", "val", "test"):
        split_df = df[df["split"] == name]
        path = args.output_dir / SPLIT_NAME.format(split=name)
        split_df.to_csv(path, index=False)
        print(f"wrote {len(split_df):>6} rows to {path}")

    linked, missing = link_imagery(df, args.dataset, args.output_dir, args.imagery)
    if args.imagery != "none":
        print(f"\n{args.imagery}ed {linked} image files into {args.output_dir}")
        if missing:
            print(f"  warning: {len(missing)} image files were not in "
                  f"{Path(args.dataset).parent}, e.g. {missing[:3]}")
    print(f"\nwrote {len(df)} rows to {args.output_dir / DATASET_NAME}")


if __name__ == "__main__":
    main()
